#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
puntenlijst_core.py
===================

Kernlogica van de Puntenlijst-generator: leest deliberatie-PDF's en bouwt
één Excel-bestand met alle examenpunten (op 20), gegroepeerd per
opleiding/afstudeerrichting.

Deze module bevat GEEN gebruikersinterface. Ze wordt gebruikt door:
  * puntenlijst_gui.py  - de grafische app (drag & drop)
  * de commandoregel:    python3 puntenlijst_core.py [map]
"""

import os
import re
import sys
import glob
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. Reguliere expressies om de tekst te ontleden
# ---------------------------------------------------------------------------

# Studenthoofding, bv: "20230477 Maes Lisa MASA - 0 Berekening: ..."
STUDENT_RE = re.compile(
    r"^(?P<nr>\d{8})\s+(?P<name>.+?)\s+(?P<status>[A-Z]{3,5})\s+-\s+\d"
)

# Planregel, bv: "Programma: M0029 - ... Plan: M0029004 - Ma VPVR: vroedvrouw spec"
PLAN_RE = re.compile(r"Plan:\s*(?P<plancode>\S+)\s*-\s*(?P<planname>.+?)\s*$")

# Vakregel-herkenning: begint met een vakcode en eindigt op
#   <PUNT> <OPLEIDINGSCODE> <aanbodsessie 4 cijfers> <volgnummer>
CODE_RE = re.compile(r"^\d{3,}[A-Z]{2,}\b")
TAIL_RE = re.compile(
    r"\s(?P<grade>\S+)\s+"
    r"(?P<tag>[A-Za-z0-9][A-Za-z0-9.\-]*)\s+"
    r"(?P<offer>\d{4})\s+\d+\s*$"
)
HEAD_RE = re.compile(
    r"^(?P<code>\S+)\s+(?P<name>.+?)\s+"
    r"(?P<sp>\d{1,2},\d{2})(?P<lect>.+?)\s+(?P<period>[SZ]\d{2})\s*$"
)

CREDIT_CODES = {"G", "VZP"}

EXAMENS_RE = re.compile(r"Totaal aantal examens:\s*(\d+)")
TEKORTEN_RE = re.compile(r"Aantal tekorten:\s*(\d+)")
RESULTAAT_RE = re.compile(r"Behaald resultaat:\s*(\d+)\s*%")
LIJST_RE = re.compile(r"\b([123])\s*-\s*(Witte|Grijze|Zwarte)\s+lijst", re.IGNORECASE)
BEOORDELING_RE = re.compile(r"\b([A-Z]{2,4})\s*-\s*([A-Z][A-Z ]{3,})\s*$")

PASS_THRESHOLD = 10  # punten < 10/20 worden gemarkeerd als tekort


# ---------------------------------------------------------------------------
# 2. Hulp: bepaal richting / jaar uit de planregel
# ---------------------------------------------------------------------------
TRACK_NAMES = {
    "M0029001": "Master - Onderzoeker in gezondheid en zorg",
    "M0029002": "Master - Leiderschap in gezondheid en zorg",
    "M0029003": "Master - Verpleegkundig specialist",
    "M0029004": "Master - Vroedvrouw specialist",
    "S0017000": "Schakeljaar - Verpleeg- en vroedkunde",
}

TRACK_SHORT = {
    "M0029001": "OGZ - Onderzoeker",
    "M0029002": "LGZ - Leiderschap",
    "M0029003": "VES - Verpleegk. spec.",
    "M0029004": "VRS - Vroedvrouw spec.",
    "S0017000": "Schakeljaar",
}


def classify_track(plancode, planname, status):
    """Geef (jaar, richtingsnaam, korte_naam) terug op basis van de planinfo."""
    plancode = (plancode or "").strip()
    if plancode.startswith("S0017") or status == "SPVP":
        jaar = "Schakeljaar"
    elif plancode.startswith("M0029") or status == "MASA":
        jaar = "Masterjaar"
    else:
        jaar = "Onbekend"
    naam = TRACK_NAMES.get(plancode)
    if not naam:
        naam = f"{jaar} - {planname}".strip(" -") if planname else (plancode or "Onbekende opleiding")
    kort = TRACK_SHORT.get(plancode) or (planname or naam)
    return jaar, naam, kort


def grade_to_value(raw):
    """Zet een ruwe puntentoken om. Geeft (weergave, numerieke_waarde_of_None)."""
    raw = raw.strip()
    if raw.isdigit():
        return str(int(raw)), int(raw)
    return raw, None


def is_deficit(disp, val):
    if val is not None:
        return val < PASS_THRESHOLD
    return disp not in CREDIT_CODES


def is_credit(disp, val):
    return val is None and disp in CREDIT_CODES


# ---------------------------------------------------------------------------
# 3. Eén PDF ontleden -> lijst van studentdicts
# ---------------------------------------------------------------------------
def parse_pdf(path, courses_registry, problems):
    students = []
    cur = None
    last_code = None
    in_courses = False

    def close():
        nonlocal cur
        if cur is not None:
            students.append(cur)
        cur = None

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line = line.rstrip()
                if not line.strip():
                    continue

                m = STUDENT_RE.match(line)
                if m:
                    close()
                    cur = {
                        "nr": m.group("nr"),
                        "naam": m.group("name").strip(),
                        "status": m.group("status"),
                        "plancode": "",
                        "planname": "",
                        "grades": {},
                        "resultaat": "",
                        "beoordeling": "",
                        "lijst": "",
                        "pdf_examens": None,
                        "pdf_tekorten": None,
                        "bron": os.path.basename(path),
                    }
                    last_code = None
                    in_courses = False
                    continue

                if cur is None:
                    continue

                mp = PLAN_RE.search(line)
                if mp and not cur["plancode"]:
                    cur["plancode"] = mp.group("plancode").strip()
                    cur["planname"] = mp.group("planname").strip()
                    continue

                if CODE_RE.match(line):
                    mt = TAIL_RE.search(line)
                    if mt:
                        code = line.split()[0]
                        disp, val = grade_to_value(mt.group("grade"))
                        if code in cur["grades"]:
                            problems.append(
                                f"{cur['bron']}: student {cur['nr']} {cur['naam']} "
                                f"heeft vak {code} meer dan één keer; laatste punt gebruikt."
                            )
                        cur["grades"][code] = (disp, val)
                        last_code = code
                        in_courses = True
                        head = line[: mt.start()]
                        mh = HEAD_RE.match(head)
                        if mh:
                            naam = mh.group("name").strip()
                            sp = mh.group("sp").replace(",", ".")
                            lect = mh.group("lect").strip()
                        else:
                            parts = head.split(maxsplit=1)
                            naam = parts[1].strip() if len(parts) > 1 else ""
                            sp, lect = "", ""
                        courses_registry.setdefault(
                            code, {"naam": naam, "sp": sp, "lector": lect}
                        )
                        continue

                if "Totaal aantal SP" in line or "Default SP vereist" in line:
                    in_courses = False
                if "Totaal aantal examens" in line:
                    me = EXAMENS_RE.search(line)
                    if me:
                        cur["pdf_examens"] = int(me.group(1))
                    mb = BEOORDELING_RE.search(line)
                    if mb:
                        cur["beoordeling"] = f"{mb.group(1)} - {mb.group(2).strip()}"
                    in_courses = False
                    continue
                if "Behaald resultaat" in line:
                    mr = RESULTAAT_RE.search(line)
                    if mr:
                        cur["resultaat"] = int(mr.group(1))
                    in_courses = False
                if "Aantal tekorten" in line:
                    mt = TEKORTEN_RE.search(line)
                    if mt:
                        cur["pdf_tekorten"] = int(mt.group(1))
                    ml = LIJST_RE.search(line)
                    if ml:
                        cur["lijst"] = f"{ml.group(1)} - {ml.group(2).capitalize()} lijst"
                    in_courses = False
                    continue
                if not cur["lijst"]:
                    ml = LIJST_RE.search(line)
                    if ml and "tekorten" not in line.lower():
                        cur["lijst"] = f"{ml.group(1)} - {ml.group(2).capitalize()} lijst"

                if in_courses and last_code and re.match(r"^[a-zà-ÿ(]", line.strip()):
                    frag = line.strip()
                    courses_registry[last_code]["naam"] += " " + frag

        close()
    return students


# ---------------------------------------------------------------------------
# 4. Controle per student
# ---------------------------------------------------------------------------
def check_student(s):
    """Geef (ok, lijst_van_meldingen)."""
    msgs = []
    found = len(s["grades"])
    fails = sum(1 for disp, v in s["grades"].values() if is_deficit(disp, v))

    if s["pdf_examens"] is not None and found != s["pdf_examens"]:
        msgs.append(
            f"aantal vakken gevonden ({found}) ≠ 'Totaal aantal examens' in PDF "
            f"({s['pdf_examens']})"
        )
    if s["pdf_tekorten"] is not None and fails != s["pdf_tekorten"]:
        msgs.append(
            f"aantal punten < {PASS_THRESHOLD} berekend ({fails}) ≠ 'Aantal tekorten' "
            f"in PDF ({s['pdf_tekorten']})"
        )
    for code, (disp, val) in s["grades"].items():
        if val is not None and not (0 <= val <= 20):
            msgs.append(f"vak {code}: punt {disp} ligt buiten 0-20")
    if not s["plancode"]:
        msgs.append("geen opleiding/plan gevonden")
    return (len(msgs) == 0), msgs


# ---------------------------------------------------------------------------
# 5. Excel opbouwen
# ---------------------------------------------------------------------------
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006", bold=True)
CREDIT_FILL = PatternFill("solid", fgColor="E2EFDA")
HEAD_FILL = PatternFill("solid", fgColor="305496")
HEAD_FONT = Font(color="FFFFFF", bold=True)
META_FILL = PatternFill("solid", fgColor="D9E1F2")
BAD_FILL = PatternFill("solid", fgColor="FFEB9C")
GOOD_FONT = Font(color="006100")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
VERT = Alignment(textRotation=90, vertical="bottom", horizontal="center", wrap_text=True)


def build_matrix_sheet(ws, students, courses_registry, show_track=False):
    META = ["Studentnr", "Naam", "Jaar"]
    if show_track:
        META.append("Richting")
    META += ["Lijst", "Resultaat %", "Eindbeoordeling"]
    nmeta = len(META)

    codes = sorted(
        {c for s in students for c in s["grades"]},
        key=lambda c: (int(re.match(r"\d+", c).group()), c),
    )

    for j, label in enumerate(META, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, CENTER, BORDER
        c2 = ws.cell(row=2, column=j, value="")
        c2.fill, c2.border = HEAD_FILL, BORDER
        ws.merge_cells(start_row=1, start_column=j, end_row=2, end_column=j)

    for k, code in enumerate(codes):
        col = nmeta + 1 + k
        reg = courses_registry.get(code, {"naam": "", "sp": "", "lector": ""})
        top = ws.cell(row=1, column=col, value=reg["naam"])
        top.fill, top.font, top.alignment, top.border = HEAD_FILL, HEAD_FONT, VERT, BORDER
        bot = ws.cell(row=2, column=col, value=code)
        bot.fill, bot.font, bot.alignment, bot.border = HEAD_FILL, HEAD_FONT, CENTER, BORDER
        comment_txt = f"{code}\n{reg['naam']}\nStudiepunten: {reg['sp']}\nLector: {reg['lector']}"
        bot.comment = Comment(comment_txt, "puntenlijst")
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    extra = ["Gem.", "# Tekorten", "Examens (PDF)", "Tekorten (PDF)", "Controle"]
    for e, label in enumerate(extra):
        col = nmeta + 1 + len(codes) + e
        c = ws.cell(row=1, column=col, value=label)
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, CENTER, BORDER
        c2 = ws.cell(row=2, column=col, value="")
        c2.fill, c2.border = HEAD_FILL, BORDER
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.row_dimensions[1].height = 150

    for i, s in enumerate(sorted(students, key=lambda x: x["naam"].lower())):
        r = 3 + i
        meta_vals = [s["nr"], s["naam"], s["status_jaar"]]
        if show_track:
            meta_vals.append(s.get("track_short", s.get("track", "")))
        meta_vals += [s["lijst"], s["resultaat"], s["beoordeling"]]
        for j, v in enumerate(meta_vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.fill, c.border = META_FILL, BORDER
            if META[j - 1] == "Resultaat %":
                c.alignment = CENTER

        numeric_vals = []
        fails = 0
        for k, code in enumerate(codes):
            col = nmeta + 1 + k
            cell = ws.cell(row=r, column=col)
            cell.alignment, cell.border = CENTER, BORDER
            if code in s["grades"]:
                disp, val = s["grades"][code]
                cell.value = val if val is not None else disp
                if val is not None:
                    numeric_vals.append(val)
                if is_deficit(disp, val):
                    fails += 1
                    cell.fill, cell.font = RED_FILL, RED_FONT
                elif is_credit(disp, val):
                    cell.fill = CREDIT_FILL
            else:
                cell.value = ""

        ok, _ = check_student(s)
        gem = round(sum(numeric_vals) / len(numeric_vals), 1) if numeric_vals else ""
        ctrl_vals = [
            gem, fails, s["pdf_examens"], s["pdf_tekorten"],
            "OK" if ok else "CONTROLEER",
        ]
        for e, v in enumerate(ctrl_vals):
            col = nmeta + 1 + len(codes) + e
            c = ws.cell(row=r, column=col, value=v)
            c.alignment, c.border = CENTER, BORDER
            if e == 4:
                if ok:
                    c.font = GOOD_FONT
                else:
                    c.fill, c.font = BAD_FILL, Font(bold=True, color="9C6500")

    META_WIDTHS = {
        "Studentnr": 12, "Naam": 26, "Jaar": 12, "Richting": 22,
        "Lijst": 16, "Resultaat %": 11, "Eindbeoordeling": 26,
    }
    for j, label in enumerate(META, start=1):
        ws.column_dimensions[get_column_letter(j)].width = META_WIDTHS.get(label, 14)

    ws.freeze_panes = ws.cell(row=3, column=nmeta + 1)


def build_legend_sheet(ws, courses_registry):
    headers = ["Vakcode", "Vaknaam", "Studiepunten", "Lector(en)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
    for i, code in enumerate(sorted(courses_registry), start=2):
        reg = courses_registry[code]
        ws.cell(row=i, column=1, value=code).border = BORDER
        ws.cell(row=i, column=2, value=reg["naam"]).border = BORDER
        ws.cell(row=i, column=3, value=reg["sp"]).border = BORDER
        ws.cell(row=i, column=4, value=reg["lector"]).border = BORDER
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "A2"


def build_check_sheet(ws, all_students):
    headers = ["Bron (PDF)", "Studentnr", "Naam", "Jaar", "Opleiding",
               "Status controle", "Detail"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
    r = 2
    n_ok = n_bad = 0
    for s in sorted(all_students, key=lambda x: (x["track"], x["naam"].lower())):
        ok, msgs = check_student(s)
        n_ok += ok
        n_bad += (not ok)
        ws.cell(row=r, column=1, value=s["bron"])
        ws.cell(row=r, column=2, value=s["nr"])
        ws.cell(row=r, column=3, value=s["naam"])
        ws.cell(row=r, column=4, value=s["status_jaar"])
        ws.cell(row=r, column=5, value=s["track"])
        status_cell = ws.cell(row=r, column=6, value="OK" if ok else "CONTROLEER")
        ws.cell(row=r, column=7, value="; ".join(msgs))
        if ok:
            status_cell.font = GOOD_FONT
        else:
            status_cell.fill, status_cell.font = BAD_FILL, Font(bold=True, color="9C6500")
        r += 1
    for col, w in zip("ABCDEFG", [28, 12, 26, 12, 42, 16, 70]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return n_ok, n_bad


def sheet_title(name, used):
    """Maak een geldige, unieke Excel-tabbladnaam (max 31 tekens)."""
    bad = r'[]:*?/\\'
    clean = "".join("-" if ch in bad else ch for ch in name)[:31]
    base = clean
    i = 2
    while clean.lower() in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


# ---------------------------------------------------------------------------
# 6. Hoofdfunctie: van PDF-paden naar Excel-bestand
# ---------------------------------------------------------------------------
def generate(pdf_paths, out_dir=None, log=print):
    """
    Verwerk een lijst PDF-bestanden en schrijf het Excel-bestand.

    pdf_paths : lijst van paden naar PDF-bestanden
    out_dir   : map waar het Excel-bestand komt (standaard: map van 1e PDF)
    log       : functie die voortgangsberichten ontvangt (bv. print)

    Geeft een dict terug met: out_path, n_students, n_ok, n_bad,
    groups (dict richting -> aantal), problems (lijst meldingen).
    """
    pdfs = sorted(pdf_paths)
    if not pdfs:
        raise ValueError("Geen PDF-bestanden opgegeven.")

    courses_registry = {}
    problems = []
    all_students = []

    for p in pdfs:
        log(f"Inlezen: {os.path.basename(p)}")
        try:
            students = parse_pdf(p, courses_registry, problems)
        except Exception as e:
            problems.append(f"Kon {os.path.basename(p)} niet verwerken: {e}")
            log(f"   !! overgeslagen door fout: {e}")
            continue
        for s in students:
            jaar, track, track_short = classify_track(
                s["plancode"], s["planname"], s["status"]
            )
            s["status_jaar"] = jaar
            s["track"] = track
            s["track_short"] = track_short
        all_students.extend(students)
        log(f"   {len(students)} studenten gevonden")

    if not all_students:
        raise ValueError(
            "Geen studenten gevonden in de PDF's. Zijn dit wel deliberatie-PDF's?"
        )

    groups = {}
    for s in all_students:
        groups.setdefault(s["track"], []).append(s)

    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()

    def group_sort_key(track):
        return (0 if track.startswith("Master") else 1 if track.startswith("Schakel") else 2, track)

    master_students = [s for s in all_students if s["status_jaar"] == "Masterjaar"]
    if master_students:
        ws = wb.create_sheet(sheet_title("Masterjaar - alle richtingen", used_titles))
        build_matrix_sheet(ws, master_students, courses_registry, show_track=True)

    for track in sorted(groups, key=group_sort_key):
        ws = wb.create_sheet(sheet_title(track, used_titles))
        build_matrix_sheet(ws, groups[track], courses_registry)

    build_check_sheet(wb.create_sheet(sheet_title("Controle", used_titles)), all_students)
    build_legend_sheet(wb.create_sheet(sheet_title("Legende", used_titles)), courses_registry)

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    out_name = f"Puntenlijst_{stamp}.xlsx"
    if out_dir is None:
        out_dir = os.path.dirname(pdfs[0])
    if not os.access(out_dir, os.W_OK):
        out_dir = os.path.expanduser("~/Desktop")
        if not os.path.isdir(out_dir) or not os.access(out_dir, os.W_OK):
            out_dir = os.getcwd()
    out_path = os.path.join(out_dir, out_name)
    wb.save(out_path)

    n_ok = sum(1 for s in all_students if check_student(s)[0])
    n_bad = len(all_students) - n_ok

    log("")
    log("=" * 50)
    log("KLAAR")
    log(f"Studenten verwerkt : {len(all_students)}")
    log(f"Opleidingen (tabbladen): {len(groups)}")
    for track in sorted(groups, key=group_sort_key):
        log(f"    - {track}: {len(groups[track])} studenten")
    log(f"Controle OK        : {n_ok}")
    log(f"Controle te checken: {n_bad}")
    if problems:
        log("")
        log("Meldingen tijdens het inlezen:")
        for pmsg in problems:
            log(f"    * {pmsg}")
    if n_bad:
        log("")
        log(">>> Bekijk het tabblad 'Controle' in het Excel-bestand.")
    log("")
    log(f"Bestand opgeslagen: {out_path}")

    return {
        "out_path": out_path,
        "n_students": len(all_students),
        "n_ok": n_ok,
        "n_bad": n_bad,
        "groups": {t: len(g) for t, g in groups.items()},
        "problems": problems,
    }


def collect_pdfs(paths):
    """Zet een mix van mappen en bestanden om in een lijst PDF-paden."""
    pdfs = []
    for p in paths:
        p = os.path.abspath(os.path.expanduser(p))
        if os.path.isdir(p):
            pdfs.extend(glob.glob(os.path.join(p, "*.pdf")))
            pdfs.extend(glob.glob(os.path.join(p, "*.PDF")))
        elif os.path.isfile(p) and p.lower().endswith(".pdf"):
            pdfs.append(p)
    # dubbels verwijderen, volgorde behouden
    seen = set()
    unique = []
    for p in pdfs:
        if p not in seen:
            seen.add(p)
            unique.append(p)
    return unique


# ---------------------------------------------------------------------------
# 7. Commandoregel-gebruik (zoals het oude script)
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) > 1:
        folder = os.path.abspath(os.path.expanduser(sys.argv[1]))
    else:
        folder = os.path.dirname(os.path.abspath(__file__))

    if not os.path.isdir(folder):
        print(f"FOUT: map bestaat niet: {folder}")
        sys.exit(1)

    pdfs = collect_pdfs([folder])
    if not pdfs:
        print(f"Geen PDF-bestanden gevonden in: {folder}")
        sys.exit(1)

    generate(pdfs, out_dir=folder)


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        print("\nER GING IETS MIS, maar het script is netjes gestopt.")
        print(f"Foutmelding: {e}")
        sys.exit(1)
